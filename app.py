"""
Sales Adjustments Filter
Filters an invoice details Excel file down to rows matching predefined sales criteria.
"""
import io
import os
from datetime import datetime
from flask import Flask, render_template, request, send_file, jsonify, abort
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024  # 200 MB total across all files

# ---------------------------------------------------------------------------
# Filter criteria — single source of truth, used by both the engine and the UI
# ---------------------------------------------------------------------------
CRITERIA = [
    {
        "id": 1,
        "label": "Grossman (YSM)",
        "description": "TextTags contains MFG",
    },
    {
        "id": 2,
        "label": "Katz",
        "description": "TextTags contains SPEC",
    },
    {
        "id": 3,
        "label": "GK",
        "sub_rules": [
            {
                "letter": "A",
                "description": "Performer/Team is Los Angeles Dodgers AND Account Email is YKRAMER@YSKG.NET AND Total Cost is $0",
            },
            {
                "letter": "B",
                "description": "Performer/Team is MLB All Star Weekend, MLB All-Star Game, or MLB Home Run Derby AND Total Cost is $0",
            },
        ],
    },
    {
        "id": 4,
        "label": "Asher (YSA, YSA 2 and YSA 3)",
        "description": "TextTags contains SCHMECK AND Total Cost is $0",
    },
    {
        "id": 5,
        "label": "TL",
        "description": "TextTags contains SPEC",
    },
    {
        "id": 6,
        "label": "Levovitz",
        "description": "TextTags contains SPEC",
    },
]

REQUIRED_COLUMNS = ["Company", "Performer/Team", "Account Email", "TextTags", "Total Cost"]


def _norm(series: pd.Series) -> pd.Series:
    """Lowercased, stripped string version of a column for case-insensitive matching."""
    return series.astype(str).str.strip().str.lower()


def _tag_contains(series: pd.Series, needle: str) -> pd.Series:
    """True where the comma-separated TextTags value contains the needle (case-insensitive)."""
    needle_lc = needle.lower()

    def check(v):
        if pd.isna(v):
            return False
        return any(p.strip().lower() == needle_lc for p in str(v).split(","))

    return series.apply(check)


def apply_filters(df: pd.DataFrame):
    """Return (filtered_df, per_rule_counts) given the raw invoice details DataFrame."""
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    company = _norm(df["Company"])
    team = _norm(df["Performer/Team"])
    email = _norm(df["Account Email"])
    tags = df["TextTags"]
    # Coerce Total Cost to numeric; non-numeric/blank → NaN (won't equal 0)
    total_cost = pd.to_numeric(df["Total Cost"], errors="coerce")

    # Rule 3 is GK LLC with two sub-rules (A: Dodgers/Kramer; B: MLB All-Star events).
    # Sub-rules combine with OR.
    rule_3a = (
        (company == "gk llc")
        & (team == "los angeles dodgers")
        & (email == "ykramer@yskg.net")
        & (total_cost == 0)
    )
    rule_3b = (
        (company == "gk llc")
        & team.isin(["mlb all star weekend", "mlb all-star game", "mlb home run derby"])
        & (total_cost == 0)
    )

    rules = {
        1: (company == "ysm tickets") & _tag_contains(tags, "mfg"),
        2: (company == "ys katz") & _tag_contains(tags, "spec"),
        3: rule_3a | rule_3b,
        4: company.isin(["ysa", "ysa 2", "ysa 3"])
        & _tag_contains(tags, "schmeck")
        & (total_cost == 0),
        5: (company == "ys tl") & _tag_contains(tags, "spec"),
        6: (company == "levovitz") & _tag_contains(tags, "spec"),
    }

    counts = {rule_id: int(mask.sum()) for rule_id, mask in rules.items()}
    combined_mask = pd.Series(False, index=df.index)
    for m in rules.values():
        combined_mask |= m

    filtered = df[combined_mask].reset_index(drop=True)
    return filtered, counts


def build_output_workbook(
    filtered: pd.DataFrame,
    counts: dict,
    source_name: str,
    per_file_stats: list = None,
) -> io.BytesIO:
    """Build the output .xlsx with a Filtered sheet and a Criteria sheet."""
    wb = Workbook()

    # ---- Filtered sheet ----
    ws = wb.active
    ws.title = "Filtered"

    # Headers
    headers = list(filtered.columns)
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="6366A8")
    header_align = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="E5E7EB")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    # Data
    for r_idx, row in enumerate(filtered.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            # Pandas can hand back NaT/NaN — normalize to None
            if pd.isna(val):
                cell.value = None
            elif isinstance(val, (pd.Timestamp, datetime)):
                cell.value = val.to_pydatetime() if isinstance(val, pd.Timestamp) else val
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            else:
                cell.value = val
            cell.border = cell_border

    ws.freeze_panes = "A2"

    # Auto-width (capped)
    for col_idx, name in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(name))
        # Sample up to 200 rows to keep this fast on large files
        sample = filtered[name].head(200) if name in filtered.columns else []
        for v in sample:
            if pd.isna(v):
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

    # ---- Criteria sheet ----
    cs = wb.create_sheet("Criteria")
    cs.column_dimensions["A"].width = 6
    cs.column_dimensions["B"].width = 30
    cs.column_dimensions["C"].width = 85
    cs.column_dimensions["D"].width = 14

    title = cs.cell(row=1, column=1, value="Filter Criteria")
    title.font = Font(name="Calibri", bold=True, size=16, color="1F2147")
    cs.merge_cells("A1:D1")

    subtitle = cs.cell(
        row=2,
        column=1,
        value=f"Source file: {source_name}    •    Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    )
    subtitle.font = Font(name="Calibri", italic=True, size=10, color="6B7280")
    cs.merge_cells("A2:D2")

    note = cs.cell(
        row=3,
        column=1,
        value="A row is kept if it matches ANY of the rules below. All comparisons are case-insensitive.",
    )
    note.font = Font(name="Calibri", size=10, color="374151")
    cs.merge_cells("A3:D3")

    header_row = 5
    for i, h in enumerate(["#", "Rule", "Definition", "Rows matched"], start=1):
        c = cs.cell(row=header_row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")

    for offset, rule in enumerate(CRITERIA, start=1):
        r = header_row + offset
        cs.cell(row=r, column=1, value=rule["id"]).alignment = Alignment(vertical="top")
        cs.cell(row=r, column=2, value=rule["label"]).alignment = Alignment(vertical="top")
        if rule.get("sub_rules"):
            desc_text = "\n".join(
                f'{sub["letter"]}. {sub["description"]}' for sub in rule["sub_rules"]
            )
        else:
            desc_text = rule["description"]
        cs.cell(row=r, column=3, value=desc_text).alignment = Alignment(
            vertical="top", wrap_text=True
        )
        cs.cell(row=r, column=4, value=counts.get(rule["id"], 0)).alignment = Alignment(
            vertical="top", horizontal="right"
        )
        # Give sub-ruled rows extra height so both lines are visible
        if rule.get("sub_rules"):
            cs.row_dimensions[r].height = 15 * (len(rule["sub_rules"]) + 1) + 4

    total_row = header_row + len(CRITERIA) + 1
    cs.cell(row=total_row, column=2, value="Total rows kept").font = Font(bold=True)
    cs.cell(row=total_row, column=4, value=len(filtered)).font = Font(bold=True)
    cs.cell(row=total_row, column=4).alignment = Alignment(horizontal="right")

    # Per-file breakdown (only shown if multiple files were merged)
    if per_file_stats and len(per_file_stats) > 1:
        section_row = total_row + 3
        title = cs.cell(row=section_row, column=1, value="Source files")
        title.font = Font(name="Calibri", bold=True, size=14, color="1F2147")
        cs.merge_cells(start_row=section_row, start_column=1, end_row=section_row, end_column=4)

        sub_header_row = section_row + 2
        for i, h in enumerate(["File", "Input rows", "Rows kept", ""], start=1):
            c = cs.cell(row=sub_header_row, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="left", vertical="center")

        for offset, stats in enumerate(per_file_stats, start=1):
            r = sub_header_row + offset
            cs.cell(row=r, column=1, value=stats["name"]).alignment = Alignment(vertical="top")
            # Span filename across columns 1-2
            cs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            cs.cell(row=r, column=3, value=stats["input_rows"]).alignment = Alignment(
                horizontal="right", vertical="top"
            )
            cs.cell(row=r, column=4, value=stats["kept_rows"]).alignment = Alignment(
                horizontal="right", vertical="top"
            )

        totals_r = sub_header_row + len(per_file_stats) + 1
        cs.cell(row=totals_r, column=1, value="Total").font = Font(bold=True)
        cs.merge_cells(start_row=totals_r, start_column=1, end_row=totals_r, end_column=2)
        total_input = sum(s["input_rows"] for s in per_file_stats)
        total_kept = sum(s["kept_rows"] for s in per_file_stats)
        c1 = cs.cell(row=totals_r, column=3, value=total_input)
        c1.font = Font(bold=True)
        c1.alignment = Alignment(horizontal="right")
        c2 = cs.cell(row=totals_r, column=4, value=total_kept)
        c2.font = Font(bold=True)
        c2.alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html", criteria=CRITERIA)


@app.route("/api/criteria")
def api_criteria():
    return jsonify(CRITERIA)


@app.route("/api/process", methods=["POST"])
def process():
    # Accept multiple files under the same field name "file" (or "files")
    files = request.files.getlist("file") or request.files.getlist("files")
    files = [f for f in files if f and f.filename]
    if not files:
        return jsonify({"error": "No file uploaded"}), 400

    for f in files:
        if not f.filename.lower().endswith((".xlsx", ".xlsm")):
            return (
                jsonify({"error": f"'{f.filename}' is not an .xlsx/.xlsm file"}),
                400,
            )

    per_file_stats = []  # [{name, input_rows, kept_rows, counts}]
    filtered_frames = []
    aggregate_counts = {rule["id"]: 0 for rule in CRITERIA}

    for f in files:
        try:
            # Each input is a single-sheet invoice details export
            df = pd.read_excel(f.stream, sheet_name=0)
        except Exception as e:
            return jsonify({"error": f"Could not read '{f.filename}': {e}"}), 400

        try:
            filtered, counts = apply_filters(df)
        except ValueError as e:
            return jsonify({"error": f"'{f.filename}': {e}"}), 400

        # Tag each row with its source file when merging — skip for single file
        if len(filtered) > 0:
            filtered = filtered.copy()
            if len(files) > 1:
                filtered.insert(0, "Source File", f.filename)
            filtered_frames.append(filtered)

        per_file_stats.append(
            {
                "name": f.filename,
                "input_rows": len(df),
                "kept_rows": len(filtered),
                "counts": counts,
            }
        )
        for rid, n in counts.items():
            aggregate_counts[rid] += n

    if filtered_frames:
        merged = pd.concat(filtered_frames, ignore_index=True)
    else:
        # All files produced zero matches — build an empty frame with the right shape
        files[0].stream.seek(0)
        empty_template = pd.read_excel(files[0].stream, sheet_name=0).iloc[0:0]
        if len(files) > 1:
            empty_template.insert(0, "Source File", pd.Series(dtype=str))
        merged = empty_template

    source_label = (
        files[0].filename
        if len(files) == 1
        else f"{len(files)} files merged"
    )
    buf = build_output_workbook(merged, aggregate_counts, source_label, per_file_stats)

    if len(files) == 1:
        base = os.path.splitext(os.path.basename(files[0].filename))[0]
        out_name = f"{base}_filtered.xlsx"
    else:
        out_name = f"merged_filtered_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=out_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/healthz")
def healthz():
    return {"status": "ok"}, 200


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port, debug=False)
